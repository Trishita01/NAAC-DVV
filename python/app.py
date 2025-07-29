import mysql.connector
from openpyxl import load_workbook
import logging
import os
import sys
from collections import defaultdict
from config import DB_CONFIG

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log')
    ]
)

def main():
    try:
        # Connect to database
        logging.info("Connecting to database...")
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # Check database connection
        if not conn.is_connected():
            raise Exception("Failed to connect to the database")
        logging.info("Successfully connected to database")

        # Fetch all 2.1.1 entries
        query = """
            SELECT year, programme_name, programme_code, no_of_seats, no_of_students
            FROM response_2_1_1
            WHERE criteria_code = '020201020101'
            ORDER BY year ASC
        """
        logging.info("Executing query...")
        cursor.execute(query)
        rows = cursor.fetchall()
        
        logging.info(f"Fetched {len(rows)} rows from database")
        
        if not rows:
            logging.warning("No data found in the database for the specified criteria")
            return
        
        logging.info(f"Sample row: {rows[0]}")

        # Group rows by year
        data_by_year = defaultdict(list)
        for row in rows:
            data_by_year[str(row['year'])].append(row)

        # Load the Excel template
        template_path = 'templates/criteria_2_1_1_template.xlsx'
        # Check if template file exists
        template_path = 'templates/criteria211Template.xlsx'
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found at: {os.path.abspath(template_path)}")
        
        logging.info("Loading Excel template...")
        wb = load_workbook(template_path)
        ws = wb.active
        logging.info("Successfully loaded Excel template")

        # Mapping year -> start row index in template
        year_to_start_row = {
            '2020': 4,   # adjust according to your sheet if you go back 5 years
            '2021': 4,
            '2022': 4,   # Year - 1 block starts at row 4
            '2023': 9,   # Year - 2 block starts at row 9
            '2024': 14,  # Year - 3 block starts at row 14
            '2025': 19,  # Year - 4 block starts at row 19
            '2026': 24   # Year - 5 block starts at row 24
        }

        # Max rows per year block
        max_rows_per_year = 3

        # Output directory
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, 'criteria_2_1_1_filled.xlsx')

        # Fill in the data
        for year, start_row in year_to_start_row.items():
            year_data = data_by_year.get(year, [])
            logging.info(f"Processing year {year} - found {len(year_data)} records")
            
            for i, record in enumerate(year_data[:max_rows_per_year]):
                row_num = start_row + i
                try:
                    ws.cell(row=row_num, column=1, value=record['programme_name'])   # Column A
                    ws.cell(row=row_num, column=2, value=record['programme_code'])   # Column B
                    ws.cell(row=row_num, column=3, value=record['no_of_seats'])      # Column C
                    ws.cell(row=row_num, column=4, value=record['no_of_students'])   # Column D
                    logging.debug(f"Filled row {row_num} with data: {record}")
                except Exception as e:
                    logging.error(f"Error filling row {row_num} with data {record}: {str(e)}")

        # Save the workbook
        logging.info(f"Saving filled template to {output_path}")
        wb.save(output_path)
        logging.info("Successfully saved the filled template")

    except mysql.connector.Error as err:
        logging.error(f"Database error: {err}")
        sys.exit(1)
    except FileNotFoundError as e:
        logging.error(f"File error: {e}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}", exc_info=True)
        sys.exit(1)
    finally:
        # Close database connection
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()
            logging.info("Database connection closed")

if __name__ == "__main__":
    main()
